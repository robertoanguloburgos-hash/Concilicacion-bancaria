# -*- coding: utf-8 -*-
"""
Conciliación Bancaria Automática — DAEM Puerto Montt
=====================================================

Aplicación Streamlit que automatiza la conciliación bancaria mensual entre:
  - Hoja 1: "1110301 MOV FONDOS"        -> Libro Mayor contable  (columnas B:M)
  - Hoja 2: "CARTOLAS BANCARIAS GENERAL" -> Cartola bancaria BCI  (columnas C:N)

Motor de conciliación en DOS FASES con soporte modular y en memoria.
"""

import io
import re
import unicodedata
from datetime import datetime, date
from itertools import combinations

import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.utils import column_index_from_string

# ---------------------------------------------------------------------------
# CONFIGURACIÓN GENERAL
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="Conciliación Bancaria Automática — DAEM", 
    page_icon="🏦",
    layout="wide"
)

SHEET_CONTABLE = "1110301 MOV FONDOS"
SHEET_BANCO = "CARTOLAS BANCARIAS GENERAL"

RANGO_CONTABLE = ("B", "M")
RANGO_BANCO = ("C", "N")

MAX_FILAS_BUSQUEDA_HEADER = 20      # filas iniciales donde se busca el encabezado real
MAX_FILAS_VACIAS_CONSECUTIVAS = 10  # corte de lectura de datos
TOLERANCIA_DIAS = 5                  # ventana +/- días para cruce por monto
TOLERANCIA_MONTO = 1.0              # margen de $1 por redondeos
MAX_COMBO_FASE2 = 6                 # tamaño máximo de combinación de abonos a probar

REGLAS_CONTABLE = {
    "Fecha": ["fecha"],
    "N°Comp.": ["comp"],
    "Glosa": ["glosa"],
    "Docum.": ["docum"],
    "DEBE": ["debe"],
    "HABER": ["haber"],
    "SALDOS": ["saldo"],
    "CRUCE CARTOLAS BANCARIAS": ["cruce", "cartola"],
}

REGLAS_BANCO = {
    "Fecha contable": ["fecha"],
    "Movimiento": ["movimiento"],
    "N° documento": ["documento"],
    "Cargo (-)": ["cargo"],
    "Abono (+)": ["abono"],
    "SALDOS": ["saldo"],
    "CARTOLA N°": ["cartola"],
    "CRUCE C/MOV FONDO": ["cruce", "mov"],
    "Glosa detalle": ["glosa"],
}

PALABRAS_CLAVE_CONTABLE = ["fecha", "comp", "glosa", "docum", "debe", "haber", "saldo", "cruce", "cartola"]
PALABRAS_CLAVE_BANCO = ["fecha", "movimiento", "documento", "cargo", "abono", "saldo", "cartola", "cruce", "glosa"]

COLUMNAS_OBLIGATORIAS_CONTABLE = ["Fecha", "DEBE", "HABER", "CRUCE CARTOLAS BANCARIAS"]
COLUMNAS_OBLIGATORIAS_BANCO = ["Fecha contable", "Cargo (-)", "Abono (+)", "CRUCE C/MOV FONDO", "CARTOLA N°"]


# ---------------------------------------------------------------------------
# UTILIDADES DE TEXTO Y PARSEO
# ---------------------------------------------------------------------------
def normalizar(texto):
    if texto is None:
        return ""
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("utf-8")
    texto = re.sub(r"\s+", " ", texto)
    return texto


def quitar_numeros(texto):
    return re.sub(r"[0-9]+", "", normalizar(texto)).strip()


def a_float(valor):
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return abs(float(valor))
    texto = str(valor).strip()
    if texto == "":
        return 0.0
    texto = texto.replace("$", "").replace(" ", "")
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        return abs(float(texto))
    except ValueError:
        return 0.0


def a_fecha(valor):
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# DETECCIÓN DINÁMICA DE ENCABEZADOS Y MAPEO DE COLUMNAS
# ---------------------------------------------------------------------------
def col_range_indices(col_ini, col_fin):
    return range(column_index_from_string(col_ini), column_index_from_string(col_fin) + 1)


def detectar_fila_header(ws, col_ini, col_fin, palabras_clave,
                         max_filas=MAX_FILAS_BUSQUEDA_HEADER, min_coincidencias=3):
    cols = list(col_range_indices(col_ini, col_fin))
    mejor_fila, mejor_score = None, 0
    for fila in range(1, max_filas + 1):
        celdas = [normalizar(ws.cell(row=fila, column=c).value) for c in cols]
        texto_fila = " | ".join(celdas)
        score = sum(1 for kw in palabras_clave if kw in texto_fila)
        if score > mejor_score:
            mejor_score, mejor_fila = score, fila
    if mejor_fila is None or mejor_score < min_coincidencias:
        return None
    return mejor_fila


def mapear_columnas(ws, fila_header, col_ini, col_fin, reglas):
    cols = list(col_range_indices(col_ini, col_fin))
    encontrados = {}
    usados = set()
    for nombre, substrings in reglas.items():
        candidato = None
        for c in cols:
            if c in usados:
                continue
            texto = normalizar(ws.cell(row=fila_header, column=c).value)
            if all(s in texto for s in substrings):
                candidato = c
                break
        encontrados[nombre] = candidato
        if candidato is not None:
            usados.add(candidato)
    return encontrados


# ---------------------------------------------------------------------------
# LECTURA DE DATOS
# ---------------------------------------------------------------------------
def leer_datos_contable(ws_valores, fila_header, colmap):
    cols = list(col_range_indices(*RANGO_CONTABLE))
    filas = []
    fila = fila_header + 1
    vacias_consecutivas = 0
    while fila <= ws_valores.max_row and vacias_consecutivas < MAX_FILAS_VACIAS_CONSECUTIVAS:
        valores = [ws_valores.cell(row=fila, column=c).value for c in cols]
        if all(v is None or str(v).strip() == "" for v in valores):
            vacias_consecutivas += 1
            fila += 1
            continue
        vacias_consecutivas = 0

        registro = {
            "fila": fila,
            "fecha": a_fecha(ws_valores.cell(row=fila, column=colmap["Fecha"]).value) if colmap.get("Fecha") else None,
            "comp": ws_valores.cell(row=fila, column=colmap["N°Comp."]).value if colmap.get("N°Comp.") else None,
            "glosa": ws_valores.cell(row=fila, column=colmap["Glosa"]).value if colmap.get("Glosa") else "",
            "docum": a_float(ws_valores.cell(row=fila, column=colmap["Docum."]).value) if colmap.get("Docum.") else 0.0,
            "debe": a_float(ws_valores.cell(row=fila, column=colmap["DEBE"]).value) if colmap.get("DEBE") else 0.0,
            "haber": a_float(ws_valores.cell(row=fila, column=colmap["HABER"]).value) if colmap.get("HABER") else 0.0,
            "usado": False,
        }
        if registro["fecha"] is not None or registro["debe"] != 0 or registro["haber"] != 0:
            filas.append(registro)
        fila += 1
    return filas


def leer_datos_banco(ws_valores, ws_escritura, fila_header, colmap):
    cols = list(col_range_indices(*RANGO_BANCO))
    filas = []
    fila = fila_header + 1
    vacias_consecutivas = 0
    cartola_actual = 1
    col_cartola = colmap.get("CARTOLA N°")

    while fila <= ws_valores.max_row and vacias_consecutivas < MAX_FILAS_VACIAS_CONSECUTIVAS:
        valores = [ws_valores.cell(row=fila, column=c).value for c in cols]
        if all(v is None or str(v).strip() == "" for v in valores):
            vacias_consecutivas += 1
            fila += 1
            continue
        vacias_consecutivas = 0

        texto_fila = normalizar(" ".join(str(v) for v in valores if v is not None))
        if "sucursal" in texto_fila or "saldo diario" in texto_fila:
            cartola_actual += 1
            fila += 1
            continue

        registro = {
            "fila": fila,
            "fecha": a_fecha(ws_valores.cell(row=fila, column=colmap["Fecha contable"]).value) if colmap.get("Fecha contable") else None,
            "movimiento": ws_valores.cell(row=fila, column=colmap["Movimiento"]).value if colmap.get("Movimiento") else "",
            "documento": a_float(ws_valores.cell(row=fila, column=colmap["N° documento"]).value) if colmap.get("N° documento") else 0.0,
            "cargo": a_float(ws_valores.cell(row=fila, column=colmap["Cargo (-)"]).value) if colmap.get("Cargo (-)") else 0.0,
            "abono": a_float(ws_valores.cell(row=fila, column=colmap["Abono (+)"]).value) if colmap.get("Abono (+)") else 0.0,
            "cartola": cartola_actual,
            "usado": False,
        }

        if col_cartola:
            ws_escritura.cell(row=fila, column=col_cartola).value = cartola_actual

        if registro["fecha"] is not None or registro["cargo"] != 0 or registro["abono"] != 0:
            filas.append(registro)

        fila += 1
    return filas


# ---------------------------------------------------------------------------
# FASE 1 — CRUCE DIRECTO (1 A 1)
# ---------------------------------------------------------------------------
def fase1_documentos(contable, banco):
    matches = []
    for c in contable:
        if c["usado"] or c["docum"] <= 0:
            continue
        for b in banco:
            if b["usado"] or b["documento"] <= 0:
                continue
            if abs(b["documento"] - c["docum"]) < 0.01:
                c["usado"] = True
                b["usado"] = True
                matches.append({"tipo": "documento", "contable": c, "bancos": [b]})
                break
    return matches


def fase1_montos(contable, banco):
    matches = []
    for c in contable:
        if c["usado"] or c["fecha"] is None:
            continue
        monto_c = c["debe"] if c["debe"] > 0 else c["haber"]
        if monto_c == 0:
            continue
        es_ingreso = c["debe"] > 0

        candidatos = []
        for b in banco:
            if b["usado"] or b["fecha"] is None:
                continue
            monto_b = b["abono"] if es_ingreso else b["cargo"]
            if monto_b <= 0:
                continue
            if abs(monto_b - monto_c) <= TOLERANCIA_MONTO:
                diff_dias = abs((b["fecha"] - c["fecha"]).days)
                if diff_dias <= TOLERANCIA_DIAS:
                    candidatos.append((diff_dias, b["fila"], b))

        if candidatos:
            candidatos.sort(key=lambda x: (x[0], x[1]))
            b_elegido = candidatos[0][2]
            c["usado"] = True
            b_elegido["usado"] = True
            matches.append({"tipo": "monto", "contable": c, "bancos": [b_elegido]})
    return matches


# ---------------------------------------------------------------------------
# FASE 2 — CONSOLIDACIÓN INTRADIARIA (1 A N)
# ---------------------------------------------------------------------------
def fase2_consolidacion(contable, banco):
    matches = []
    huerfanos = [c for c in contable if not c["usado"] and c["debe"] > 0 and c["fecha"] is not None]

    for c in huerfanos:
        candidatos = [b for b in banco if not b["usado"] and b["abono"] > 0 and b["fecha"] == c["fecha"]]
        if not candidatos:
            continue

        grupos = {}
        for b in candidatos:
            clave = quitar_numeros(b["movimiento"])
            grupos.setdefault(clave, []).append(b)

        encontrado = False

        for grupo in grupos.values():
            suma = sum(g["abono"] for g in grupo)
            if len(grupo) > 1 and abs(suma - c["debe"]) <= TOLERANCIA_MONTO:
                for g in grupo:
                    g["usado"] = True
                c["usado"] = True
                matches.append({"tipo": "consolidado", "contable": c, "bancos": grupo})
                encontrado = True
                break
        if encontrado:
            continue

        for grupo in grupos.values():
            if len(grupo) < 2:
                continue
            combo_elegido = None
            for tam in range(2, min(len(grupo), MAX_COMBO_FASE2) + 1):
                for combo in combinations(grupo, tam):
                    suma = sum(g["abono"] for g in combo)
                    if abs(suma - c["debe"]) <= TOLERANCIA_MONTO:
                        combo_elegido = combo
                        break
                if combo_elegido:
                    break
            if combo_elegido:
                for g in combo_elegido:
                    g["usado"] = True
                c["usado"] = True
                matches.append({"tipo": "consolidado", "contable": c, "bancos": list(combo_elegido)})
                break

    return matches


# ---------------------------------------------------------------------------
# ESCRITURA DE RESULTADOS
# ---------------------------------------------------------------------------
def escribir_resultados(ws_contable, colmap_c, ws_banco, colmap_b, matches, contable_all, banco_all):
    col_cruce_c = colmap_c.get("CRUCE CARTOLAS BANCARIAS")
    col_cruce_b = colmap_b.get("CRUCE C/MOV FONDO")
    col_glosa_b = colmap_b.get("Glosa detalle")

    for m in matches:
        c = m["contable"]
        bancos = m["bancos"]
        monto_banco_total = sum(abs(b["abono"] if b["abono"] > 0 else b["cargo"]) for b in bancos)
        comp_texto = str(c["comp"]) if c["comp"] is not None else ""
        monto_contable = c["debe"] if c["debe"] > 0 else c["haber"]

        if col_cruce_c:
            if m["tipo"] == "consolidado":
                valor = f"Consolidado x{len(bancos)} (Cartola {bancos[0]['cartola']})"
            else:
                valor = monto_banco_total
            ws_contable.cell(row=c["fila"], column=col_cruce_c).value = valor

        for b in bancos:
            if col_cruce_b:
                ws_banco.cell(row=b["fila"], column=col_cruce_b).value = monto_contable
            if col_glosa_b:
                ws_banco.cell(row=b["fila"], column=col_glosa_b).value = comp_texto

    if col_cruce_c:
        for c in contable_all:
            if not c["usado"] and (c["debe"] != 0 or c["haber"] != 0):
                ws_contable.cell(row=c["fila"], column=col_cruce_c).value = "PENDIENTE"

    if col_cruce_b:
        for b in banco_all:
            if not b["usado"] and (b["cargo"] != 0 or b["abono"] != 0):
                ws_banco.cell(row=b["fila"], column=col_cruce_b).value = "PENDIENTE"


def procesar_archivo(bytes_entrada):
    wb_escritura = openpyxl.load_workbook(io.BytesIO(bytes_entrada), data_only=False)
    wb_valores = openpyxl.load_workbook(io.BytesIO(bytes_entrada), data_only=True)

    for nombre_hoja in (SHEET_CONTABLE, SHEET_BANCO):
        if nombre_hoja not in wb_escritura.sheetnames:
            raise ValueError(f"No se encontró la hoja '{nombre_hoja}' en el archivo.")

    ws_contable_w = wb_escritura[SHEET_CONTABLE]
    ws_banco_w = wb_escritura[SHEET_BANCO]
    ws_contable_v = wb_valores[SHEET_CONTABLE]
    ws_banco_v = wb_valores[SHEET_BANCO]

    fila_header_c = detectar_fila_header(ws_contable_v, *RANGO_CONTABLE, PALABRAS_CLAVE_CONTABLE)
    fila_header_b = detectar_fila_header(ws_banco_v, *RANGO_BANCO, PALABRAS_CLAVE_BANCO)

    if fila_header_c is None:
        raise ValueError(f"No se detectó el encabezado en '{SHEET_CONTABLE}' (columnas {RANGO_CONTABLE[0]}:{RANGO_CONTABLE[1]}).")
    if fila_header_b is None:
        raise ValueError(f"No se detectó el encabezado en '{SHEET_BANCO}' (columnas {RANGO_BANCO[0]}:{RANGO_BANCO[1]}).")

    colmap_c = mapear_columnas(ws_contable_v, fila_header_c, *RANGO_CONTABLE, REGLAS_CONTABLE)
    colmap_b = mapear_columnas(ws_banco_v, fila_header_b, *RANGO_BANCO, REGLAS_BANCO)

    faltantes_c = [n for n in COLUMNAS_OBLIGATORIAS_CONTABLE if colmap_c.get(n) is None]
    faltantes_b = [n for n in COLUMNAS_OBLIGATORIAS_BANCO if colmap_b.get(n) is None]
    if faltantes_c:
        raise ValueError(f"Faltan columnas en la hoja contable: {', '.join(faltantes_c)}.")
    if faltantes_b:
        raise ValueError(f"Faltan columnas en la hoja bancaria: {', '.join(faltantes_b)}.")

    datos_contable = leer_datos_contable(ws_contable_v, fila_header_c, colmap_c)
    datos_banco = leer_datos_banco(ws_banco_v, ws_banco_w, fila_header_b, colmap_b)

    m1a = fase1_documentos(datos_contable, datos_banco)
    m1b = fase1_montos(datos_contable, datos_banco)
    m2 = fase2_consolidacion(datos_contable, datos_banco)
    todos_los_matches = m1a + m1b + m2

    escribir_resultados(ws_contable_w, colmap_c, ws_banco_w, colmap_b, todos_los_matches, datos_contable, datos_banco)

    salida = io.BytesIO()
    wb_escritura.save(salida)
    salida.seek(0)

    resumen = {
        "total_contable": len(datos_contable),
        "total_banco": len(datos_banco),
        "cruces_documento": len(m1a),
        "cruces_monto": len(m1b),
        "cruces_consolidados": len(m2),
        "pendientes_contable": sum(1 for c in datos_contable if not c["usado"] and (c["debe"] != 0 or c["haber"] != 0)),
        "pendientes_banco": sum(1 for b in datos_banco if not b["usado"] and (b["cargo"] != 0 or b["abono"] != 0)),
        "matches": todos_los_matches,
    }
    return salida, resumen


# INTERFAZ STREAMLIT
st.title("🏦 Conciliación Bancaria Automática — DAEM Puerto Montt")
st.caption(
    "Cruce híbrido entre el Libro Mayor contable (1110301 MOV FONDOS) y la Cartola BCI "
    "(CARTOLAS BANCARIAS GENERAL), en dos fases: cruce directo y consolidación intradiaria."
)

with st.expander("ℹ️ Cómo funciona esta herramienta"):
    st.markdown(
        """
- **Fase 1 — Cruce directo (1 a 1):** primero empareja por **N° de documento** exacto
  (cheques y transferencias identificadas). Lo que no calza por documento se empareja
  por **monto exacto** (DEBE↔Abono, HABER↔Cargo) dentro de una ventana de **±5 días**.
- **Fase 2 — Consolidación intradiaria (1 a N):** para ingresos contables (DEBE) que
  quedaron sin cruzar, agrupa los abonos bancarios **del mismo día** con descriptores
  de **Movimiento** similares, y si la suma exacta del grupo calza con el monto
  consolidado, cruza todo el grupo de una vez.
- Los **saldos intermedios no se usan como criterio de validación** (el orden de las
  transacciones del día difiere entre contabilidad y banco); solo se usan como
  referencia secundaria para desempate.
- Cada vez que la cartola bancaria indica **"SUCURSAL"** o **"SALDO DIARIO"**, se
  detecta un nuevo folio y se incrementa automáticamente el número de **CARTOLA N°**.
- Todo el procesamiento ocurre **100% en memoria** (nunca se escribe en disco) y se
  **preservan fórmulas, formatos y colores** originales del Excel.
        """
    )

archivo = st.file_uploader("Sube el archivo Excel mensual (.xlsx)", type=["xlsx"])

if archivo is not None:
    if st.button("🚀 Ejecutar conciliación", type="primary"):
        with st.spinner("Procesando conciliación en memoria..."):
            try:
                bytes_entrada = archivo.read()
                salida, resumen = procesar_archivo(bytes_entrada)
            except Exception as e:
                st.error(f"❌ Ocurrió un error al procesar el archivo: {e}")
                st.stop()

        st.success("✅ Conciliación completada correctamente.")

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Movimientos contables", resumen["total_contable"])
        c2.metric("Movimientos bancarios", resumen["total_banco"])
        c3.metric("Cruces Fase 1", resumen["cruces_documento"] + resumen["cruces_monto"])
        c4.metric("Cruces Fase 2 (consolidados)", resumen["cruces_consolidados"])

        c5, c6, c7 = st.columns(3)
        c5.metric("↳ por N° documento", resumen["cruces_documento"])
        c6.metric("Pendientes Libro Mayor", resumen["pendientes_contable"])
        c7.metric("Pendientes Cartola", resumen["pendientes_banco"])

        detalle = []
        for m in resumen["matches"]:
            c = m["contable"]
            for b in m["bancos"]:
                detalle.append(
                    {
                        "Tipo de cruce": m["tipo"],
                        "Fila contable": c["fila"],
                        "Fecha contable": c["fecha"],
                        "Comprobante": c["comp"],
                        "Monto contable": c["debe"] if c["debe"] > 0 else c["haber"],
                        "Fila banco": b["fila"],
                        "Fecha banco": b["fecha"],
                        "Monto banco": b["abono"] if b["abono"] > 0 else b["cargo"],
                        "Cartola N°": b["cartola"],
                    }
                )

        if detalle:
            st.subheader("Detalle de cruces realizados")
            df_detalle = pd.DataFrame(detalle).sort_values(["Fila contable", "Fila banco"])
            st.dataframe(df_detalle, use_container_width=True, hide_index=True)

        st.download_button(
            label="⬇️ Descargar Excel conciliado",
            data=salida,
            file_name="conciliacion_bancaria_resultado.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
else:
    st.info(
        "Esperando la carga del archivo Excel mensual con las hojas "
        f"'{SHEET_CONTABLE}' y '{SHEET_BANCO}'."
    )
